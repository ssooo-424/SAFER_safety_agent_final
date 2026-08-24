# noqa: SIZE_OK -- Excel 원본을 OpenAI API로 처리해 pilot workbook을 만드는 단일 실험 CLI이므로 분리하지 않습니다.
from __future__ import annotations

import os
import re
import json
import time
from pathlib import Path
from typing import List, Dict, Any, Optional

from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


VERSION = "AI_SCENARIO_LINKED_REFINED_V4"

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
ENV_PATH = PROJECT_ROOT / ".env"

load_dotenv(ENV_PATH)

INPUT = SCRIPT_DIR / "risk_assessment_normalized.xlsx"
OUTPUT = SCRIPT_DIR / "감소대책_시나리오연결_파일럿_30건.xlsx"

MODEL = os.getenv("OPENAI_MODEL", "gpt-4.1-mini")
PILOT_N = 30
MAX_MEASURES = 8

client: Optional[OpenAI] = None


# WHY: 전체 원본을 API에 보내기 전에 다양한 text pattern을 포함한 후보를 고르기 위해 rule-based 1차 selection을 사용합니다.

STRONG_SPLIT_RE = re.compile(r"(?:\r?\n+|;\s*|(?<=[다함됨음요])\.\s+)")
NUMBER_BULLET_RE = re.compile(r"(?m)^\s*(?:[①②③④⑤⑥⑦⑧⑨⑩]|\d+[.)]|[-•·])\s*")

MANAGER_WORDS = [
    "교육", "배치", "지급", "관리감독", "작업계획서",
    "허가서", "출입통제", "통제", "설치", "협의",
    "점검 실시", "비치", "수립", "작성", "설정",
    "방호", "안전인증"
]

WORKER_WORDS = [
    "착용", "체결", "확인", "이동", "접근 금지",
    "접근금지", "사용 금지", "사용금지", "작업 중지",
    "신호에 따라", "정리정돈", "제거", "준수", "사용"
]

PPE_HINTS = [
    "안전모", "안전대", "안전화", "보안경",
    "보안면", "장갑", "마스크", "보호구",
    "절연용 보호구"
]


def normalize_protected_terms(text: str) -> str:
    """
    슬래시가 있어도 하나의 표현으로 유지해야 하는 단어를 보호합니다.
    예: 직 / 간접 접촉 금지 → 직·간접 접촉 금지
    """
    replacements = [
        (r"직\s*[/·ㆍ]\s*간접", "직·간접"),
        (r"유\s*/\s*무", "유·무"),
        (r"상\s*/\s*하", "상·하"),
        (r"전\s*/\s*후", "전·후"),
    ]

    for pattern, replacement in replacements:
        text = re.sub(pattern, replacement, text)

    return text


def normalize(text: str) -> str:
    text = (text or "").replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n[ \t]+", "\n", text)
    text = normalize_protected_terms(text)
    return text.strip()


def first_pass_split(text: str) -> List[str]:
    """
    1차 분리 함수입니다.

    주의:
    - 쉼표는 자동 분리하지 않습니다.
    - 슬래시는 자동 분리하지 않습니다.
    - 줄바꿈, 세미콜론, 문장 종결 마침표만 강하게 분리합니다.
    """
    text = normalize(text)

    if not text:
        return []

    text = NUMBER_BULLET_RE.sub("\n", text)

    parts = [
        part.strip(" .")
        for part in STRONG_SPLIT_RE.split(text)
        if part and part.strip(" .")
    ]

    return parts


def classify_features(text: str) -> Dict[str, object]:
    t = normalize(text)

    manager = any(word in t for word in MANAGER_WORDS)
    worker = any(word in t for word in WORKER_WORDS)

    if manager and worker:
        role = "혼합"
    elif manager:
        role = "관리자/시설 중심"
    elif worker:
        role = "작업자 중심"
    else:
        role = "판단 필요"

    features = []

    if "\n" in t:
        features.append("줄바꿈")
    if "/" in (text or "") or "·" in t:
        features.append("슬래시/복합어")
    if ";" in t:
        features.append("세미콜론")
    if "." in t:
        features.append("마침표")
    if t.count(",") >= 2:
        features.append("쉼표 다수")
    if "(" in t and ")" in t and any(x in t for x in PPE_HINTS):
        features.append("보호구 목록")
    if manager:
        features.append("관리자 동사")
    if worker:
        features.append("작업자 동사")
    if len(t) > 120:
        features.append("장문")
    elif len(t) < 35:
        features.append("단문")

    return {
        "role": role,
        "features": ", ".join(features) or "일반",
        "length": len(t),
    }


def choose_representative_rows(ws, n: int = 30):
    rows = []

    for row_num in range(2, ws.max_row + 1):
        text = ws.cell(row_num, 11).value or ""
        features = classify_features(text)
        parts = first_pass_split(text)

        rows.append({
            "row": row_num,
            "text": text,
            "parts": parts,
            **features,
        })

    desired_tags = [
        "줄바꿈",
        "슬래시/복합어",
        "마침표",
        "쉼표 다수",
        "보호구 목록",
        "관리자 동사",
        "작업자 동사",
        "장문",
        "단문",
    ]

    selected = []
    used_rows = set()

    for tag in desired_tags:
        candidates = [
            row for row in rows
            if tag in row["features"] and row["row"] not in used_rows
        ]

        candidates.sort(key=lambda row: (-len(row["parts"]), -row["length"]))

        for candidate in candidates[:3]:
            selected.append(candidate)
            used_rows.add(candidate["row"])

            if len(selected) >= n:
                break

        if len(selected) >= n:
            break

    for role in ["혼합", "관리자/시설 중심", "작업자 중심", "판단 필요"]:
        candidates = [
            row for row in rows
            if row["role"] == role and row["row"] not in used_rows
        ]

        candidates.sort(key=lambda row: (-len(row["parts"]), -row["length"]))

        for candidate in candidates[:3]:
            selected.append(candidate)
            used_rows.add(candidate["row"])

            if len(selected) >= n:
                break

        if len(selected) >= n:
            break

    if len(selected) < n:
        step = max(1, len(rows) // (n - len(selected) + 1))

        for i in range(0, len(rows), step):
            candidate = rows[i]

            if candidate["row"] not in used_rows:
                selected.append(candidate)
                used_rows.add(candidate["row"])

            if len(selected) >= n:
                break

    return sorted(selected[:n], key=lambda row: row["row"])


SYSTEM_PROMPT = """
너는 건설 안전교육을 20년 이상 진행해 온 건설안전관리 전문가다.

너는 건설현장에서 실제 작업자들을 대상으로 안전교육을 진행하기 위해,
위험 상황별 감소대책을 교육용 안전수칙 문장으로 정제하는 역할을 맡고 있다.

이 작업의 목적은 작업자들이 감소대책을 보고
'어떤 위험시나리오에서 무엇을 지켜야 사고 위험을 막을 수 있는지' 이해할 수 있도록
원문 감소대책을 명확하고 실행 가능한 한국어 안전대책 문장으로 다듬는 것이다.

현재 단계의 작업은 최종 개인화 수칙을 만드는 것이 아니다.
즉, 성향별 문장 생성이나 트리거별 개인화 문장을 만드는 단계가 아니다.

너의 작업은 원문 감소대책을 의미 단위로 분리하고,
각 대책을 위험시나리오와 연결된 서술형 한국어 안전대책 문장으로 정제하는 것이다.

핵심 원칙:
- 원문 감소대책의 의미를 보존한다.
- 원문에 없는 새로운 장비, 작업명, 위험요인을 추가하지 않는다.
- 관리자/시설 조치가 원문에 있으면 그 조치를 없애지 말고 교육용 안전대책 문장으로 정제한다.
- 작업자가 이해할 수 있도록 문장을 자연스럽게 만들되, 원문 조치의 주체와 의미를 임의로 바꾸지 않는다.
- 최종 문장에는 가능한 한 위험시나리오의 작업 상황과 사고 위험이 함께 드러나야 한다.

문장 구조:
각 대책 문장은 가능하면 다음 구조를 따른다.

[위험시나리오의 작업 상황]에서
[사고로 이어지는 위험을 막기 위해]
[감소대책을 수행한다].

예:
- 위험시나리오: 슬립폼 하부 작업대에서 견출 작업 중 공구 또는 자재가 외부로 낙하되어 아래 작업자가 맞음
- 원문 감소대책: 방호선반 설치
- 정제문: 슬립폼 하부 작업대에서 견출 작업을 할 때 공구나 자재가 외부로 낙하하지 않도록 방호선반을 설치한다.

세부 기준:

1. 명사형 감소대책은 서술어 문장으로 변환하되, 가능하면 위험시나리오 맥락을 반영한다.
- 발파작업 금지사항 준수 → 발파작업 시 사고를 막기 위해 발파작업 금지사항을 준수한다.
- 제작장에 소화기등 비치 → 제작장에서 화재 발생에 대비해 소화기등을 비치한다.
- 소화설비 설치 → 화재 위험이 있는 작업구역에는 소화설비를 설치한다.
- 방호선반 설치 → 작업 중 자재나 공구가 낙하하지 않도록 방호선반을 설치한다.
- 출입금지구역 설정 → 낙하 또는 충돌 위험이 있는 작업구역에는 출입금지구역을 설정한다.
- 적합한 절연용 보호구 착용 → 전기 접촉 위험이 있는 작업 시 적합한 절연용 보호구를 착용한다.
- 절연용 방호구 설치 → 전기 접촉 위험을 막기 위해 절연용 방호구를 설치한다.
- 활선작업용 기구 및 장치 사용 → 활선작업 시 감전 위험을 막기 위해 활선작업용 기구 및 장치를 사용한다.

2. 명사만 있는 경우에는 사고 맥락을 보고 역할을 추론해 문장으로 구성한다.
단, 명사 단독 대책은 '설치한다'로 단정하지 말고 설치 상태 확인 문장으로 만든다.
- 낙하물방지망 → 고소작업 또는 낙하 위험 작업 시 낙하물방지망 설치 상태를 확인한다.
- 안전난간 → 추락 위험이 있는 작업 전 안전난간 설치 상태를 확인한다.
- 개구부 덮개 → 개구부 주변 작업 전 개구부 덮개 설치 상태를 확인한다.

3. 명사 단독 대책과 명사+동작명사 대책을 반드시 구분한다.
- 낙하물방지망 → 낙하물방지망 설치 상태를 확인한다.
- 방호선반 설치 → 방호선반을 설치한다.
- 출입금지구역 설정 → 출입금지구역을 설정한다.
- 소화설비 설치 → 소화설비를 설치한다.
- 보호구 착용 → 보호구를 착용한다.

4. 사고내용을 반영해 작업자가 이해할 수 있는 교육용 안전대책 문장으로 정제한다.
단순히 '방호선반을 설치한다'처럼 짧게 쓰지 말고,
가능하면 '어떤 작업 상황에서 어떤 사고 위험을 막기 위해 방호선반을 설치하는지'가 드러나게 작성한다.

예:
- 원문: 방호선반 설치
  사고내용: 슬립폼 하부 작업대에서 견출 작업 중 공구 또는 자재가 외부로 낙하되어 아래 작업자가 맞음
  정제문: 슬립폼 하부 작업대에서 견출 작업을 할 때 공구나 자재가 외부로 낙하하지 않도록 방호선반을 설치한다.

- 원문: 출입금지구역 설정
  사고내용: 공구 또는 자재가 외부로 낙하되어 아래 작업자가 맞음
  정제문: 슬립폼 하부 작업대에서 견출 작업을 할 때 낙하 위험 구역에 출입금지구역을 설정한다.

- 원문: 낙하물방지망
  사고내용: 공구 또는 자재가 외부로 낙하되어 아래 작업자가 맞음
  정제문: 슬립폼 하부 작업대에서 견출 작업을 할 때 낙하물방지망 설치 상태를 확인한다.

- 원문: 소화설비 설치
  사고내용: 화기 사용 또는 화재 위험
  정제문: 화기 사용 작업을 할 때 화재 발생에 대비해 소화설비를 설치한다.

- 원문: 안전대 착용
  사고내용: 추락 위험이 있는 장소에서 작업
  정제문: 추락 위험이 있는 장소에서 작업할 때 안전대를 착용한다.

5. 직 / 간접, 유 / 무, 전 / 후 같은 표현은 쪼개지 않는다.
- 직 / 간접 접촉 금지 → 직·간접 접촉을 금지한다.

6. 접근한계거리 이내 접근 같은 표현은 금지 문장으로 바꾼다.
- 접근한계거리 이내 접근 → 접근한계거리 이내로 접근하지 않는다.

7. 보호구 목록은 하나의 문장으로 유지한다.
- 안전모, 안전대, 안전화 착용 → 안전모, 안전대, 안전화를 착용한다.
- 적합한 절연용 보호구 착용 → 적합한 절연용 보호구를 착용한다.

8. 하위 항목이 독립 대책이면 분리한다.
예:
장비 설치작업시 작업계획서 작성
- 설치 대상에 대한 정보 확인
- 설비의 설치순서 및 방법
- 전도방지 조치 등 보강조치 수립

정제:
- 장비 설치작업 시 작업계획서를 작성한다.
- 장비 설치 전 설치 대상에 대한 정보를 확인한다.
- 장비 설치 전 설비의 설치순서 및 방법을 확인한다.
- 장비 전도 위험을 막기 위해 전도방지 조치 등 보강조치를 수립한다.

9. 원문에 없는 작업자 행동으로 과하게 바꾸지 않는다.
현재 단계에서는 작업자 개인화 행동 변환이 아니라, 감소대책 정제 단계다.
- 신호수 배치 → 신호수를 배치한다.
- 작업계획서 작성 → 작업계획서를 작성한다.
- 방호선반 설치 → 방호선반을 설치한다.
다만 위험시나리오 맥락을 넣어 교육용 문장으로 자연스럽게 확장할 수 있다.

10. 트리거와 안전성향은 현재 단계에서 문장에 넣지 않는다.
예를 들어 '일정에 쫓기더라도', '편한 방식으로 넘기지 말고' 같은 표현은 아직 쓰지 않는다.
이 표현은 이후 최종 수칙 생성 단계에서 별도 라이브러리로 반영한다.

출력은 반드시 JSON 객체 하나만 반환한다.

출력 JSON 형식:
{
  "measure_count": number,
  "measures": [
    {
      "raw_segment": "원문에서 대응되는 구간",
      "scenario_linked_sentence": "위험시나리오와 연결된 서술형 안전대책 문장",
      "role": "작업자 중심 | 관리자/시설 중심 | 혼합 | 판단 필요",
      "reason": "분리 및 문장화 판단 근거"
    }
  ],
  "review_needed": "Y | N",
  "review_note": "검수가 필요한 이유. 없으면 빈 문자열"
}
"""


def call_ai_for_measure(payload: Dict[str, Any]) -> Dict[str, Any]:
    global client

    if client is None:
        raise RuntimeError("OpenAI client가 초기화되지 않았습니다.")

    user_prompt = f"""
다음 감소대책을 정제해줘.

[위험시나리오 및 사고 맥락]
- 대공정: {payload.get("major_process", "")}
- 공정내용: {payload.get("process_content", "")}
- 세부공정: {payload.get("detail_process", "")}
- 설비: {payload.get("equipment", "")}
- 물질: {payload.get("material", "")}
- 위험시나리오/유해위험요인: {payload.get("risk_scenario", "")}
- 사고분류: {payload.get("accident_type", "")}
- 관련근거: {payload.get("legal_basis", "")}

[감소대책 원문]
{payload.get("raw_measure", "")}

[규칙 기반 1차 후보]
{json.dumps(payload.get("first_pass_candidates", []), ensure_ascii=False)}

위 1차 후보는 참고용일 뿐이다.
잘못 쪼개졌으면 다시 합치고, 빠진 분리가 있으면 새로 분리해라.

반드시 지켜야 할 것:
- 최종 문장은 위험시나리오와 연결된 교육용 안전대책 문장이어야 한다.
- 단순히 '방호선반을 설치한다.'처럼 짧게 끝내지 말고, 가능하면 어떤 사고 위험을 막기 위한 조치인지 드러내라.
- 다만 원문에 없는 장비, 작업명, 위험요인을 새로 만들지 마라.
- '낙하물방지망'처럼 명사만 있으면 '설치한다'로 만들지 말고, 사고 맥락을 반영해 '설치 상태를 확인한다'로 만든다.
- '방호선반 설치', '출입금지구역 설정'처럼 명사+동작명사 구조는 해당 동작을 서술어로 바꾸되, 위험시나리오 맥락을 반영한다.
- '직 / 간접 접촉 금지'를 '직'과 '간접 접촉 금지'로 나누지 마라.
- 보호구 목록은 하나의 착용 문장으로 유지해라.
- 트리거 표현이나 성향 표현은 넣지 마라. 예: '일정에 쫓기더라도', '편한 방식으로 넘기지 말고' 같은 표현 금지.

반드시 JSON만 출력해라.
"""

    response = client.chat.completions.create(
        model=MODEL,
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": user_prompt},
        ],
        response_format={"type": "json_object"},
        temperature=0.1,
    )

    content = response.choices[0].message.content or ""

    try:
        data = json.loads(content)
    except json.JSONDecodeError:
        return {
            "measure_count": 0,
            "measures": [],
            "review_needed": "Y",
            "review_note": f"JSON 파싱 실패: {content[:300]}",
        }

    if "measures" not in data or not isinstance(data["measures"], list):
        return {
            "measure_count": 0,
            "measures": [],
            "review_needed": "Y",
            "review_note": "AI 응답에 measures 배열이 없음",
        }

    cleaned_measures = []

    for measure in data["measures"]:
        if not isinstance(measure, dict):
            continue

        raw_segment = str(measure.get("raw_segment", "") or "").strip()
        sentence = str(measure.get("scenario_linked_sentence", "") or "").strip()
        role = str(measure.get("role", "") or "").strip()
        reason = str(measure.get("reason", "") or "").strip()

        if sentence and not sentence.endswith("."):
            sentence += "."

        if role not in ["작업자 중심", "관리자/시설 중심", "혼합", "판단 필요"]:
            role = "판단 필요"

        cleaned_measures.append({
            "raw_segment": raw_segment,
            "scenario_linked_sentence": sentence,
            "role": role,
            "reason": reason,
        })

    data["measures"] = cleaned_measures
    data["measure_count"] = len(cleaned_measures)

    review_needed = str(data.get("review_needed", "N") or "N").strip().upper()
    data["review_needed"] = "Y" if review_needed == "Y" else "N"
    data["review_note"] = str(data.get("review_note", "") or "").strip()

    if data["measure_count"] == 0:
        data["review_needed"] = "Y"
        data["review_note"] = data["review_note"] or "분리된 대책이 없음"

    if data["measure_count"] > MAX_MEASURES:
        data["review_needed"] = "Y"
        data["review_note"] = (
            data["review_note"] + f" / 대책 수가 {MAX_MEASURES}개를 초과해 엑셀에는 앞 {MAX_MEASURES}개만 표시됨"
        ).strip(" /")

    return data


def style_sheet(ws, freeze: str = "A2"):
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E1F2")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin_gray)

    ws.auto_filter.ref = ws.dimensions


def get_measure_value(ai_data: Dict[str, Any], index: int, key: str) -> str:
    measures = ai_data.get("measures", [])

    if index >= len(measures):
        return ""

    return str(measures[index].get(key, "") or "")


def reason_summary(ai_data: Dict[str, Any]) -> str:
    reasons = []

    for measure in ai_data.get("measures", []):
        reason = measure.get("reason", "")

        if reason:
            reasons.append(reason)

    return " / ".join(reasons[:4])


def role_summary(ai_data: Dict[str, Any]) -> str:
    roles = []

    for measure in ai_data.get("measures", []):
        role = measure.get("role", "")

        if role and role not in roles:
            roles.append(role)

    if not roles:
        return "판단 필요"

    if len(roles) == 1:
        return roles[0]

    return "혼합"


def build_error_ai_data(error_message: str) -> Dict[str, Any]:
    return {
        "measure_count": 0,
        "measures": [],
        "review_needed": "Y",
        "review_note": error_message,
    }


def main():
    global client

    print(f"실행 중인 코드 버전: {VERSION}")
    print(f"입력 파일: {INPUT}")
    print(f"출력 파일: {OUTPUT}")
    print(f"환경 파일: {ENV_PATH}")
    print(f"사용 모델: {MODEL}")
    print()

    api_key = os.getenv("OPENAI_API_KEY")

    if not api_key:
        raise RuntimeError(
            f"OPENAI_API_KEY가 설정되어 있지 않습니다. "
            f".env 파일을 확인하세요: {ENV_PATH}"
        )

    client = OpenAI(api_key=api_key)

    if not INPUT.exists():
        raise FileNotFoundError(f"입력 파일을 찾을 수 없습니다: {INPUT}")

    src_wb = load_workbook(INPUT, data_only=False)
    src_ws = src_wb[src_wb.sheetnames[0]]
    selected = choose_representative_rows(src_ws, PILOT_N)

    wb = Workbook()

    guide = wb.active
    guide.title = "안내"
    guide.sheet_view.showGridLines = False

    guide["A1"] = "감소대책 시나리오 연결 파일럿 사용법"
    guide["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    guide["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    guide.merge_cells("A1:F1")

    notes = [
        ("목적", "전체 237건 처리 전에 30건으로 위험시나리오와 연결된 기본 안전대책 문장 생성 기준을 확인합니다."),
        ("처리 방식", "규칙 기반 1차 후보를 만든 뒤 OpenAI API가 의미 단위로 다시 분리하고, 위험시나리오와 연결된 교육용 안전대책 문장으로 정제합니다."),
        ("중요", "현재 단계에서는 트리거 표현과 안전성향 표현을 넣지 않습니다. 이것은 이후 최종 수칙 생성 단계에서 별도 라이브러리로 반영합니다."),
        ("검수 방법", "파일럿_30건 시트에서 'AI 정제 대책 1~8' 열을 먼저 확인합니다."),
        ("좋은 결과 예시", "슬립폼 하부 작업대에서 견출 작업을 할 때 공구나 자재가 외부로 낙하하지 않도록 방호선반을 설치한다."),
        ("수정 필요 예시", "방호선반을 설치한다. → 시나리오 연결이 부족함"),
        ("다음 단계", "파일럿 결과가 괜찮으면 같은 방식으로 전체 237건을 처리해 measure_library.xlsx/json을 만듭니다."),
        ("사용 모델", MODEL),
        ("환경파일", str(ENV_PATH)),
    ]

    for i, (key, value) in enumerate(notes, 3):
        guide.cell(i, 1, key).font = Font(bold=True, color="666666")
        guide.cell(i, 2, value).alignment = Alignment(wrap_text=True, vertical="top")

    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 120

    ws = wb.create_sheet("파일럿_30건")

    source_headers = [
        src_ws.cell(1, col).value
        for col in range(1, src_ws.max_column + 1)
    ]

    extra_headers = [
        "원본행번호",
        "AI 최종 대책 수",
    ]

    for i in range(1, MAX_MEASURES + 1):
        extra_headers.append(f"AI 정제 대책 {i}")

    for i in range(1, MAX_MEASURES + 1):
        extra_headers.append(f"AI 원문구간 {i}")

    for i in range(1, MAX_MEASURES + 1):
        extra_headers.append(f"AI 조치성격 {i}")

    extra_headers += [
        "AI 주체 요약",
        "AI 판단 근거",
        "검수 필요",
        "검수 메모",
        "패턴태그",
        "1차 주체추정",
        "1차 후보수",
        "1차 후보 전체",
    ]

    headers = source_headers + extra_headers
    ws.append(headers)

    for idx, item in enumerate(selected, start=1):
        row_num = item["row"]

        source_values = [
            src_ws.cell(row_num, col).value
            for col in range(1, src_ws.max_column + 1)
        ]

        raw_measure = src_ws.cell(row_num, 11).value or ""
        first_candidates = first_pass_split(raw_measure)

        payload = {
            "major_process": src_ws.cell(row_num, 1).value or "",
            "process_content": src_ws.cell(row_num, 2).value or "",
            "detail_process": src_ws.cell(row_num, 3).value or "",
            "equipment": src_ws.cell(row_num, 4).value or "",
            "material": src_ws.cell(row_num, 5).value or "",
            "risk_scenario": src_ws.cell(row_num, 6).value or "",
            "accident_type": src_ws.cell(row_num, 7).value or "",
            "legal_basis": src_ws.cell(row_num, 12).value or "",
            "raw_measure": raw_measure,
            "first_pass_candidates": first_candidates,
        }

        print(f"[{idx}/{len(selected)}] AI 처리 중 - 원본행 {row_num}")

        try:
            ai_data = call_ai_for_measure(payload)
        except Exception as error:
            ai_data = build_error_ai_data(f"API 오류: {error}")

        row = source_values + [
            row_num,
            ai_data.get("measure_count", 0),
        ]

        for measure_index in range(MAX_MEASURES):
            row.append(get_measure_value(ai_data, measure_index, "scenario_linked_sentence"))

        for measure_index in range(MAX_MEASURES):
            row.append(get_measure_value(ai_data, measure_index, "raw_segment"))

        for measure_index in range(MAX_MEASURES):
            row.append(get_measure_value(ai_data, measure_index, "role"))

        row += [
            role_summary(ai_data),
            reason_summary(ai_data),
            ai_data.get("review_needed", "Y"),
            ai_data.get("review_note", ""),
            item["features"],
            item["role"],
            len(first_candidates),
            "\n".join(first_candidates),
        ]

        ws.append(row)

        # WHY: OpenAI API 요청 사이에 짧은 간격을 두어 pilot 실행의 pacing을 유지합니다.
        time.sleep(0.2)

    style_sheet(ws)

    ws.row_dimensions[1].height = 42

    source_col_count = len(source_headers)
    ai_start_col = source_col_count + 1

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if cell.column <= source_col_count:
                cell.font = Font(color="008000")
            elif cell.column <= source_col_count + 2 + MAX_MEASURES:
                cell.font = Font(color="0000FF")
            elif cell.column <= source_col_count + 2 + (MAX_MEASURES * 3):
                cell.font = Font(color="808080")
            else:
                cell.font = Font(color="0000FF")

    for row in ws.iter_rows(min_row=2, min_col=ai_start_col, max_col=source_col_count + 2 + MAX_MEASURES):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="FFF2CC")

    header_to_col = {
        ws.cell(1, col).value: col
        for col in range(1, ws.max_column + 1)
    }

    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(1, col).value or "")

        if header == "감소대책":
            width = 70
        elif header == "유해위험요인":
            width = 60
        elif header.startswith("AI 정제 대책"):
            width = 65
        elif header.startswith("AI 원문구간"):
            width = 34
        elif header.startswith("AI 조치성격"):
            width = 22
        elif header in ["AI 판단 근거", "검수 메모"]:
            width = 55
        elif header == "1차 후보 전체":
            width = 45
        elif header in ["패턴태그", "1차 주체추정"]:
            width = 28
        elif col <= source_col_count:
            width = 18
        else:
            width = 20

        ws.column_dimensions[get_column_letter(col)].width = width

    for row_num in range(2, ws.max_row + 1):
        ws.row_dimensions[row_num].height = 130

    dv_yn = DataValidation(
        type="list",
        formula1='"Y,N"',
        allow_blank=True,
    )

    ws.add_data_validation(dv_yn)

    review_col = get_column_letter(header_to_col["검수 필요"])
    dv_yn.add(f"{review_col}2:{review_col}{ws.max_row}")

    rules = wb.create_sheet("분리기준")

    rules.append(["우선순위", "판단 기준", "AI 처리 원칙", "예시"])

    rule_rows = [
        (
            1,
            "위험시나리오 연결",
            "단순 문장화가 아니라 작업 상황과 사고 위험을 포함한다.",
            "방호선반 설치 → 슬립폼 하부 작업대에서 견출 작업을 할 때 공구나 자재가 외부로 낙하하지 않도록 방호선반을 설치한다.",
        ),
        (
            2,
            "명사 단독",
            "설치한다로 단정하지 않고 설치 상태 확인 문장으로 만든다.",
            "낙하물방지망 → 슬립폼 하부 작업대에서 견출 작업을 할 때 낙하물방지망 설치 상태를 확인한다.",
        ),
        (
            3,
            "명사+동작명사",
            "동작명사를 서술어로 바꾸되, 위험시나리오 맥락을 반영한다.",
            "출입금지구역 설정 → 낙하 위험 구역에 출입금지구역을 설정한다.",
        ),
        (
            4,
            "쉼표",
            "보호구·장비 목록이면 분리하지 않는다.",
            "안전모, 안전대, 안전화 착용 → 안전모, 안전대, 안전화를 착용한다.",
        ),
        (
            5,
            "슬래시",
            "하나의 복합어면 분리하지 않는다.",
            "직 / 간접 접촉 금지 → 직·간접 접촉을 금지한다.",
        ),
        (
            6,
            "하위 항목",
            "독립 대책이면 분리한다.",
            "작업계획서 작성 / 정보 확인 / 설치순서 확인 / 보강조치 수립은 각각 분리 가능",
        ),
        (
            7,
            "트리거 제외",
            "현재 단계에서는 트리거 표현을 넣지 않는다.",
            "일정에 쫓기더라도, 익숙하더라도, 불편하더라도 같은 표현은 이후 단계에서 처리",
        ),
        (
            8,
            "성향 제외",
            "현재 단계에서는 성향별 취약 행동 표현을 넣지 않는다.",
            "편한 방식으로 넘기지 말고, 본인 발밑을 놓치지 말고 등은 이후 단계에서 처리",
        ),
    ]

    for rule in rule_rows:
        rules.append(rule)

    style_sheet(rules)

    rules.column_dimensions["A"].width = 12
    rules.column_dimensions["B"].width = 28
    rules.column_dimensions["C"].width = 55
    rules.column_dimensions["D"].width = 95

    for row in rules.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # WHY: input과 같은 reference_data 디렉터리에 output을 저장해 pilot 결과의 provenance를 유지합니다.
    wb.save(OUTPUT)

    print()
    print("완료:", OUTPUT.name)
    print("저장 위치:", OUTPUT.resolve())
    print("selected source rows:", [item["row"] for item in selected])


if __name__ == "__main__":
    main()
